"""
QA API Test Runner
Reads a QA test case xlsx, executes API test cases, updates results.

Usage:
    python run_api_tests.py <xlsx_path> <base_url> [options]

Options:
    --auth-method login|bearer|none  (default: none)
    --token <bearer_token>           (required if auth-method=bearer)
    --email <email>                  (required if auth-method=login)
    --password <password>            (required if auth-method=login)
    --login-endpoint <path>          (default: /api/v1/auth/login)
    --token-field <field>            (default: access_token)
    --filter category:API|priority:P1|range:TC-001:TC-010|all  (default: all)
    --timeout <seconds>              (default: 30)
    --output <output_path>           (default: overwrites input file)
    --verify-ssl                     (default: True, use --no-verify-ssl to disable)
"""

import argparse
import json
import re
import sys
import time
from datetime import datetime

import requests
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# === STYLES ===
PASS_FILL = PatternFill('solid', fgColor='C6EFCE')
FAIL_FILL = PatternFill('solid', fgColor='FFC7CE')
SKIP_FILL = PatternFill('solid', fgColor='FFF9C4')
DATA_FONT = Font(name='Arial', size=10)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def parse_test_data(text):
    """Parse structured Test Data column into a dict."""
    result = {}
    if not text:
        return result
    for line in str(text).strip().split('\n'):
        line = line.strip()
        if not line or line.startswith('#'):
            continue
        if ': ' not in line:
            continue
        key, value = line.split(': ', 1)
        key = key.strip().upper()
        try:
            result[key] = json.loads(value)
        except (json.JSONDecodeError, ValueError):
            result[key] = value.strip()
    return result


def is_api_testable(parsed_data):
    """Check if a test case has the minimum required fields for API execution."""
    return all(k in parsed_data for k in ['METHOD', 'ENDPOINT', 'EXPECT_STATUS'])


def substitute_variables(text, variables):
    """Replace {{variable}} placeholders with actual values."""
    if not isinstance(text, str):
        return text
    for key, val in variables.items():
        text = text.replace(f'{{{{{key}}}}}', str(val))
    # Handle nested response references: {{tc_001_response.field}}
    pattern = r'\{\{(\w+)\.(.+?)\}\}'
    for match in re.finditer(pattern, text):
        var_name, field_path = match.group(1), match.group(2)
        if var_name in variables and isinstance(variables[var_name], dict):
            obj = variables[var_name]
            for part in field_path.split('.'):
                if isinstance(obj, dict) and part in obj:
                    obj = obj[part]
                else:
                    obj = None
                    break
            if obj is not None:
                text = text.replace(match.group(0), str(obj))
    return text


def deep_substitute(obj, variables):
    """Recursively substitute variables in dicts, lists, and strings."""
    if isinstance(obj, str):
        return substitute_variables(obj, variables)
    elif isinstance(obj, dict):
        return {k: deep_substitute(v, variables) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [deep_substitute(item, variables) for item in obj]
    return obj


def execute_test(parsed_data, base_url, token, variables, timeout=30, verify_ssl=True):
    """Execute a single API test case. Returns (pass/fail, detail_message)."""
    method = parsed_data['METHOD'].upper()
    endpoint = substitute_variables(str(parsed_data['ENDPOINT']), variables)
    url = f"{base_url.rstrip('/')}{endpoint}"

    expect_status = int(parsed_data['EXPECT_STATUS'])

    headers = {"Content-Type": "application/json"}
    if token:
        headers["Authorization"] = f"Bearer {token}"

    extra_headers = parsed_data.get('HEADERS', {})
    if isinstance(extra_headers, dict):
        extra_headers = deep_substitute(extra_headers, variables)
        # If Authorization is explicitly set (even empty), override
        if 'Authorization' in extra_headers:
            if extra_headers['Authorization'] == '':
                del headers['Authorization']
            else:
                headers['Authorization'] = extra_headers['Authorization']
        headers.update({k: v for k, v in extra_headers.items() if k != 'Authorization'})

    body = parsed_data.get('BODY')
    if body:
        body = deep_substitute(body, variables)

    query = parsed_data.get('QUERY')
    if query:
        query = deep_substitute(query, variables)

    try:
        response = requests.request(
            method, url,
            json=body if body else None,
            params=query if query else None,
            headers=headers,
            timeout=timeout,
            verify=verify_ssl
        )
    except requests.Timeout:
        return False, f"Timeout after {timeout}s", None
    except requests.ConnectionError as e:
        return False, f"Connection error: {str(e)[:100]}", None
    except Exception as e:
        return False, f"Request error: {str(e)[:100]}", None

    # Parse response
    resp_body = None
    try:
        resp_body = response.json()
    except (json.JSONDecodeError, ValueError):
        resp_body = response.text[:200] if response.text else None

    failures = []

    # Check status code
    if response.status_code != expect_status:
        failures.append(f"Status: expected {expect_status}, got {response.status_code}")

    # Check body contains keys
    expect_contains = parsed_data.get('EXPECT_BODY_CONTAINS')
    if expect_contains and isinstance(resp_body, dict):
        for key in expect_contains:
            if key not in resp_body:
                failures.append(f"Missing key: '{key}'")

    # Check body match
    expect_match = parsed_data.get('EXPECT_BODY_MATCH')
    if expect_match and isinstance(resp_body, dict):
        for key, val in expect_match.items():
            actual = resp_body.get(key)
            if actual != val:
                failures.append(f"Mismatch '{key}': expected {val}, got {actual}")

    # Check body not contains
    expect_not = parsed_data.get('EXPECT_BODY_NOT_CONTAINS')
    if expect_not and isinstance(resp_body, dict):
        for key in expect_not:
            if key in resp_body:
                failures.append(f"Unexpected key: '{key}'")

    # Check error message
    expect_err = parsed_data.get('EXPECT_ERROR_MESSAGE')
    if expect_err:
        resp_text = json.dumps(resp_body) if isinstance(resp_body, dict) else str(resp_body)
        if expect_err not in resp_text:
            failures.append(f"Error message '{expect_err}' not found in response")

    # Check array length
    if isinstance(resp_body, dict):
        data_field = resp_body.get('data', resp_body)
        if isinstance(data_field, list):
            min_len = parsed_data.get('EXPECT_ARRAY_MIN_LENGTH')
            if min_len is not None and len(data_field) < int(min_len):
                failures.append(f"Array length {len(data_field)} < min {min_len}")
            max_len = parsed_data.get('EXPECT_ARRAY_MAX_LENGTH')
            if max_len is not None and len(data_field) > int(max_len):
                failures.append(f"Array length {len(data_field)} > max {max_len}")

    if failures:
        detail = f"Status={response.status_code}. FAILED: {'; '.join(failures)}"
        return False, detail[:200], resp_body
    else:
        detail = f"Status={response.status_code} OK"
        if isinstance(resp_body, dict):
            snippet = json.dumps(resp_body)[:120] + ('...' if len(json.dumps(resp_body)) > 120 else '')
            detail += f" | {snippet}"
        return True, detail[:200], resp_body


def authenticate(base_url, method, token=None, email=None, password=None,
                 login_endpoint='/api/v1/auth/login', token_field='access_token',
                 verify_ssl=True):
    """Obtain bearer token based on auth method."""
    if method == 'none':
        return None, None
    if method == 'bearer':
        return token, None
    if method == 'login':
        url = f"{base_url.rstrip('/')}{login_endpoint}"
        try:
            resp = requests.post(url, json={"email": email, "password": password},
                                 timeout=30, verify=verify_ssl)
            if resp.status_code >= 400:
                return None, f"Auth failed: status {resp.status_code} — {resp.text[:200]}"
            data = resp.json()
            # Try common token field names
            tok = data.get(token_field)
            if not tok:
                for field in ['access_token', 'token', 'data.token', 'data.access_token']:
                    parts = field.split('.')
                    obj = data
                    for p in parts:
                        obj = obj.get(p, {}) if isinstance(obj, dict) else None
                    if obj and isinstance(obj, str):
                        tok = obj
                        break
            if not tok:
                return None, f"Auth succeeded but token not found. Response: {json.dumps(data)[:200]}"
            return tok, None
        except Exception as e:
            return None, f"Auth error: {str(e)[:200]}"
    return None, f"Unknown auth method: {method}"


def should_run(tc_id, priority, category, filter_str):
    """Determine if a test case matches the user's filter."""
    if filter_str == 'all':
        return True
    if filter_str.startswith('category:'):
        return category and filter_str.split(':', 1)[1].lower() in category.lower()
    if filter_str.startswith('priority:'):
        return priority and filter_str.split(':', 1)[1].upper() in priority.upper()
    if filter_str.startswith('range:'):
        parts = filter_str.split(':', 1)[1].split(':')
        if len(parts) == 2:
            start_num = int(re.search(r'\d+', parts[0]).group())
            end_num = int(re.search(r'\d+', parts[1]).group())
            tc_num = int(re.search(r'\d+', tc_id).group()) if re.search(r'\d+', tc_id) else 0
            return start_num <= tc_num <= end_num
    return True


def run_tests(xlsx_path, base_url, auth_method='none', token=None, email=None,
              password=None, login_endpoint='/api/v1/auth/login', token_field='access_token',
              filter_str='all', timeout=30, output_path=None, verify_ssl=True):
    """Main test runner."""
    output_path = output_path or xlsx_path
    start_time = time.time()

    # Load workbook
    wb = load_workbook(xlsx_path)
    ws = wb['Test Cases']

    # Authenticate
    print(f"[AUTH] Authenticating with method: {auth_method}")
    bearer_token, auth_err = authenticate(
        base_url, auth_method, token, email, password, login_endpoint, token_field, verify_ssl
    )
    if auth_err:
        print(f"[AUTH FAIL] {auth_err}")
        sys.exit(1)
    print(f"[AUTH] {'Token obtained' if bearer_token else 'No auth'}")

    # Runtime variables
    variables = {
        'base_url': base_url,
        'token': bearer_token or '',
        'timestamp': datetime.now().isoformat(),
    }

    stats = {'total': 0, 'executed': 0, 'passed': 0, 'failed': 0, 'skipped': 0}
    results_log = []

    # Iterate test cases (row 2 onwards)
    for row in range(2, ws.max_row + 1):
        tc_id = ws[f'A{row}'].value
        if not tc_id:
            continue
        stats['total'] += 1

        module = ws[f'B{row}'].value or ''
        category = ws[f'C{row}'].value or ''
        priority = ws[f'D{row}'].value or ''
        test_data_raw = ws[f'H{row}'].value or ''

        # Check filter
        if not should_run(tc_id, priority, category, filter_str):
            continue

        # Parse test data
        parsed = parse_test_data(test_data_raw)

        if not is_api_testable(parsed):
            ws[f'J{row}'] = 'Skipped'
            ws[f'J{row}'].fill = SKIP_FILL
            ws[f'J{row}'].font = DATA_FONT
            ws[f'J{row}'].border = THIN_BORDER
            ws[f'L{row}'] = f"Manual test — requires human execution"
            ws[f'L{row}'].font = DATA_FONT
            ws[f'L{row}'].border = THIN_BORDER
            stats['skipped'] += 1
            results_log.append({'tc_id': tc_id, 'status': 'Skipped', 'detail': 'Not API-parseable'})
            continue

        # Execute
        print(f"[RUN] {tc_id}: {parsed['METHOD']} {parsed['ENDPOINT']}")
        passed, detail, resp_body = execute_test(
            parsed, base_url, bearer_token, variables, timeout, verify_ssl
        )
        stats['executed'] += 1

        # Save response for chaining
        save_as = parsed.get('SAVE_RESPONSE')
        if save_as and resp_body:
            variables[save_as] = resp_body

        # Write results
        status = 'Pass' if passed else 'Fail'
        if passed:
            stats['passed'] += 1
            ws[f'J{row}'].fill = PASS_FILL
        else:
            stats['failed'] += 1
            ws[f'J{row}'].fill = FAIL_FILL

        ws[f'J{row}'] = status
        ws[f'J{row}'].font = DATA_FONT
        ws[f'J{row}'].border = THIN_BORDER

        ws[f'K{row}'] = 'Claude AI (automated)'
        ws[f'K{row}'].font = DATA_FONT
        ws[f'K{row}'].border = THIN_BORDER

        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        existing_notes = ws[f'L{row}'].value or ''
        ws[f'L{row}'] = f"[{timestamp}] {detail}" + (f" | {existing_notes}" if existing_notes else "")
        ws[f'L{row}'].font = DATA_FONT
        ws[f'L{row}'].border = THIN_BORDER
        ws[f'L{row}'].alignment = Alignment(wrap_text=True, vertical='top')

        print(f"  → {status}: {detail[:80]}")
        results_log.append({'tc_id': tc_id, 'status': status, 'detail': detail})

    # === ADD TEST RUN LOG SHEET ===
    elapsed = round(time.time() - start_time, 2)

    if 'Test Run Log' in wb.sheetnames:
        del wb['Test Run Log']
    log_ws = wb.create_sheet('Test Run Log')

    log_ws.column_dimensions['A'].width = 25
    log_ws.column_dimensions['B'].width = 50

    log_data = [
        ('Test Run Summary', ''),
        ('Run Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        ('Base URL', base_url[:30] + '***' if len(base_url) > 30 else base_url),
        ('Auth Method', auth_method),
        ('Filter', filter_str),
        ('Duration', f'{elapsed}s'),
        ('', ''),
        ('Results', ''),
        ('Total Test Cases', stats['total']),
        ('Executed', stats['executed']),
        ('Passed', stats['passed']),
        ('Failed', stats['failed']),
        ('Skipped', stats['skipped']),
        ('Pass Rate', f"{round(stats['passed']/max(stats['executed'],1)*100, 1)}%"),
    ]

    LABEL_FONT = Font(name='Arial', size=10, bold=True)
    TITLE_FONT = Font(name='Arial', size=12, bold=True)

    for i, (label, value) in enumerate(log_data, 1):
        log_ws[f'A{i}'] = label
        log_ws[f'B{i}'] = value
        if label in ('Test Run Summary', 'Results'):
            log_ws[f'A{i}'].font = TITLE_FONT
        else:
            log_ws[f'A{i}'].font = LABEL_FONT
            log_ws[f'B{i}'].font = DATA_FONT

    # Save
    wb.save(output_path)
    print(f"\n[DONE] Results saved to {output_path}")
    print(f"  Total={stats['total']} Executed={stats['executed']} "
          f"Passed={stats['passed']} Failed={stats['failed']} Skipped={stats['skipped']}")
    return stats


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='QA API Test Runner')
    parser.add_argument('xlsx_path', help='Path to QA test case xlsx')
    parser.add_argument('base_url', help='Base URL of the API')
    parser.add_argument('--auth-method', default='none', choices=['login', 'bearer', 'none'])
    parser.add_argument('--token', default=None)
    parser.add_argument('--email', default=None)
    parser.add_argument('--password', default=None)
    parser.add_argument('--login-endpoint', default='/api/v1/auth/login')
    parser.add_argument('--token-field', default='access_token')
    parser.add_argument('--filter', default='all', dest='filter_str')
    parser.add_argument('--timeout', type=int, default=30)
    parser.add_argument('--output', default=None)
    parser.add_argument('--no-verify-ssl', action='store_true')

    args = parser.parse_args()
    run_tests(
        args.xlsx_path, args.base_url,
        auth_method=args.auth_method, token=args.token,
        email=args.email, password=args.password,
        login_endpoint=args.login_endpoint, token_field=args.token_field,
        filter_str=args.filter_str, timeout=args.timeout,
        output_path=args.output, verify_ssl=not args.no_verify_ssl
    )
